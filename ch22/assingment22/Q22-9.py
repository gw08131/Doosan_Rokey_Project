# Q22-9.py

from openpyxl import Workbook, load_workbook


path = "C:/rokey/python/ch22/assingment22/data.xlsx"

try:
    wb = load_workbook(path)
except FileNotFoundError:
    # 워크북 생성
    wb = Workbook()
    wb.save(path)

    