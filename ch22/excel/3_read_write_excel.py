# 3_read_write_excel.py

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

def readExcel(sheet,max_row,max_col):
    for row in sheet.iter_rows(max_row = max_row, max_col=max_col):
        for cell in row:
            print(cell.value, end="\t")
        print()


def create_exel_file(file_path,data):
    """
    새로운 엑셀 파일을 생성하고 기본 데이터를 입력
    param = file_path: 생성할 엑셀 파일 경로
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Data"

    # ws.append(["이름","나이","국가"])
    for row in data:
        ws.append(row)
    wb.save(file_path)
    print(f"엑셀 파일 생성 완료: {file_path}")
    
    return ws


if __name__ == "__main__":
    data_list = [
        ["이름","나이","국가"],
        ["홍길동",30,"대한민국"],
        ["Alice",25,"USA"],
        ["Bob",28,"UK"]
                ]
    excel_path = r"C:\rokey\python\ch22\excel/sample.xlsx"
    sheet = create_exel_file(excel_path,data_list)

    readExcel(sheet,4,3)