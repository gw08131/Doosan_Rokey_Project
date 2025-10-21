# openpyxl_module.py

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# book > sheet > cell

# book 객체 생성 및 오픈
wb0 = Workbook()     # 새로운 워크북 생성
sheet0 = wb0.active   # 활성화 시트 선택

# 셀에 데이터 입력
sheet0['A1'] = 'Hello'
sheet0['B1'] = 'World'

# 엑셀 파일 저장
file =r'C:\rokey\python\ch22\excel\new_example0.xlsx'
wb0.save(file)

print("--------------------")
# Excel 파일 열기 (book 불러와서 생성/읽기)
file1 =r'C:\rokey\python\ch22\excel\new_example1.xlsx'
wb1 = Workbook() 
wb1.save(file1)
wb1 = load_workbook(file1)
sheet1 = wb1.active
sheet1['A1'] = 'Hello'
sheet1['B1'] = 'My'
sheet1['C1'] = 'name'
sheet1['D1'] = 'is'
sheet1['E1'] = 'Jiyun'

print(sheet1['A1'].value)
print(sheet1['B1'].value)
print(sheet1['C2'].value)
file1 =r'C:\rokey\python\ch22\excel\new_example1.xlsx'
wb1.save(file1)

print("--------------------")
# 여러 셀의 데이터 읽기
wb2 = Workbook()
sheet2 = wb2.active
for row in sheet2.iter_rows(min_row=1,max_row=5,max_col=8):
    for cell in row:
        print(cell.value)

print("--------------------")
def readExcel(max_row,max_col):
    for row in sheet2.iter_rows(max_row = max_row, max_col=max_col):
        for cell in row:
            print(cell.value, end="\t")
        print()

readExcel(3,5)

print("--------------------")
# Excel 파일 수정: 셀 값 수정하기
# A1 셀값 수정
sheet0['A1'] = "새로운 값"
file =r'C:\rokey\python\ch22\excel\new_example0.xlsx'
wb0.save(file)

readExcel(2,3)

print("--------------------")
# 다양한 데이터 형식 다루기
# 숫자, 날짜, 수식 입력 (실행결과는 엑셀을 열어 확인)

from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()     # 새로운 워크북 생성
sheet = wb.active

# 날짜 입력
sheet['A1'] = datetime(2025,10,15)
# 숫자 입력
sheet['B1'] = 100
# 수식 입력
sheet['C1'] = '=B1*2'

# 저장하기
file =r'C:\rokey\python\ch22\excel\formuals_example.xlsx'
wb.save(file)

readExcel(3,3)

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(3))